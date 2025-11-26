#!/usr/bin/env python3
"""
app_optimized.py - High-performance async email verifier (fixed DNS handling)

This file merges the full validation & scoring logic of the older 'app.py'
with the performance improvements from the optimized variant.

NOTES (fixes applied):
- Central DNS_TIMEOUT/DNS_LIFETIME used consistently in _resolver()
- lookup_mx_async properly implemented (no double-await, clear exception handling)
- DNSTimeoutError used to indicate hard DNS timeouts (caller marks invalid)
- validate_single_async uses asyncio.create_task + wait_for and catches DNSTimeoutError
"""

from __future__ import annotations
import os
import re
import csv
import json
import time
import random
import socket
import logging
import asyncio
import tempfile
import smtplib
import hashlib
import urllib.request
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict
from datetime import datetime

import dns.resolver
import dns.exception
from dns.resolver import LifetimeTimeout, NoNameservers

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ==================== CONFIG ====================

LOG_LEVEL = logging.INFO

# DNS settings (centralized)
DNS_TIMEOUT = float(os.getenv("DNS_TIMEOUT", "3"))        # per query socket timeout (seconds)
DNS_LIFETIME = float(os.getenv("DNS_LIFETIME", "9"))      # total lifetime for resolver operations
MAX_DNS_RETRIES = int(os.getenv("MAX_DNS_RETRIES", "1"))  # how many resolver attempts before hard timeout
DOMAIN_CACHE_TTL = int(os.getenv("DOMAIN_CACHE_TTL", "3600"))

# SMTP settings - choose mode via env or CLI
SMTP_MODE = os.getenv("SMTP_MODE", "balanced")  # "fast", "balanced", "thorough"
SMTP_TIMEOUTS = {
    "fast": 3,      # 3s - fast but less thorough
    "balanced": 5,  # default
    "thorough": 10  # thorough but slow
}
SMTP_TIMEOUT = SMTP_TIMEOUTS.get(SMTP_MODE, 5)
SMTP_RETRY_COUNT = int(os.getenv("SMTP_RETRY_COUNT", "1"))
SMTP_POOL_MAX = int(os.getenv("SMTP_POOL_MAX", "50"))

# Worker / concurrency
CPU_CORES = os.cpu_count() or 4
DEFAULT_WORKERS = max(4, CPU_CORES * 4)
MAX_CONCURRENT_SMTP = max(4, CPU_CORES * 8)

# Thread pool for blocking libraries (dnspython/smtplib/urllib)
_EXEC = ThreadPoolExecutor(max_workers=min(128, CPU_CORES * 16))

logger = logging.getLogger("verifier")
_handler = logging.StreamHandler()
_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(_handler)
logger.setLevel(LOG_LEVEL)

class DNSTimeoutError(Exception):
    """Raised when DNS lookup consistently times out for a domain."""

# ==================== REGEXES & CONSTANTS ====================

EMAIL_REGEX = re.compile(
    r'^(?:"[^"]+"|[A-Za-z0-9!#$%&\'*+/=?^_`{|}~.-]+)@'
    r'(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?\.)+'
    r'[A-Za-z]{2,}$'
)
EMAIL_TOKEN_SPLIT = re.compile(r"[,\s;/|]+")
EMAIL_STRIP_CHARS = " <>\"'()[]{}:;,"

SMTP_USER_NOT_FOUND_RE = re.compile(
    r"user not found|unknown user|no such user|recipient .* not found|mailbox unavailable|"
    r"unknown recipient|550 5\.1\.1|recipient unknown|no mailbox here by that name",
    re.IGNORECASE,
)

ROLE_LOCALPARTS = {
    "admin","administrator","info","support","sales","contact","help","office","postmaster"
}
FREE_PROVIDERS = {"gmail.com","yahoo.com","outlook.com","hotmail.com","icloud.com","aol.com"}

PROVIDER_RULES = {
    "gmail.com": {"mode": "trust_mx"},
    "googlemail.com": {"mode": "trust_mx"},
    "hotmail.com": {"mode": "trust_mx"},
    "outlook.com": {"mode": "trust_mx"},
    "live.com": {"mode": "trust_mx"},
    "yahoo.com": {"mode": "strict"},
    "aol.com": {"mode": "strict"},
    "gmx.com": {"mode": "strict"},
    "zoho.com": {"mode": "strict"},
}

DISPOSABLE_FALLBACK = {
    "mailinator.com","tempmail.net","10minutemail.com","trashmail.com","guerrillamail.com","yopmail.com"
}
DEFAULT_SOURCES = [
    "https://raw.githubusercontent.com/disposable/disposable-email-domains/master/disposable_email_blacklist.conf",
    "https://disposable.github.io/disposable-email-domains/domains.json",
]

DISPOSABLE_MX_KEYWORDS = {
    "mailinator","yopmail","guerrillamail","trashmail","tempmail","10minutemail","temp-mail"
}
DISPOSABLE_ASN_LIST = {
    a.strip().upper() for a in (os.getenv("DISPOSABLE_ASN_LIST", "").split(",")) if a.strip()
}

# ==================== DYNAMIC DISPOSABLE DOMAINS ====================

def _fetch_text(url: str, timeout: int = 6) -> Optional[str]:
    try:
        with urllib.request.urlopen(url, timeout=timeout) as r:
            data = r.read()
            try:
                return data.decode("utf-8")
            except Exception:
                return data.decode("latin-1", errors="ignore")
    except Exception:
        return None

def _parse_domains(payload: str) -> List[str]:
    payload = (payload or "").strip()
    if not payload:
        return []
    try:
        arr = json.loads(payload)
        if isinstance(arr, list):
            return [str(x).strip().lower().rstrip(".") for x in arr if isinstance(x, str) and x.strip()]
    except Exception:
        pass
    out = []
    for line in payload.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or " " in line:
            continue
        out.append(line.lower().rstrip("."))
    return out

def load_disposable_domains_dynamic() -> set:
    disable_fetch = os.getenv("DISPOSABLE_DISABLE_FETCH", "false").lower() == "true"
    ttl = int(os.getenv("DISPOSABLE_TTL_SECONDS", "86400"))
    cache_path = os.getenv("DISPOSABLE_CACHE_PATH") or os.path.join(tempfile.gettempdir(), "disposable_domains.cache")
    src_env = os.getenv("DISPOSABLE_SOURCES")
    sources = [s.strip() for s in src_env.split(",")] if src_env else DEFAULT_SOURCES

    try:
        if os.path.exists(cache_path):
            if (time.time() - os.path.getmtime(cache_path)) <= ttl:
                with open(cache_path, "r", encoding="utf-8") as f:
                    arr = json.load(f)
                    if isinstance(arr, list):
                        return set([d.strip().lower() for d in arr if isinstance(d, str) and d.strip()])
    except Exception:
        pass

    merged = set()
    if not disable_fetch:
        for url in sources:
            txt = _fetch_text(url, timeout=6)
            if not txt:
                continue
            merged.update(_parse_domains(txt))
        if merged:
            try:
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(sorted(list(merged)), f)
            except Exception:
                pass
            return merged
    return set(DISPOSABLE_FALLBACK)

DISPOSABLE_DOMAINS = load_disposable_domains_dynamic()

# ==================== TTL Cache (sync and async helpers) ====================

class TTLCacheSync:
    def __init__(self, ttl_seconds: int = 3600, max_items: int = 20000):
        self.ttl = ttl_seconds
        self.max_items = max_items
        self._store: Dict[str, Tuple[float, Any]] = {}

    def get(self, key: str) -> Optional[Any]:
        v = self._store.get(key)
        if not v:
            return None
        ts, data = v
        if (time.time() - ts) > self.ttl:
            self._store.pop(key, None)
            return None
        return data

    def set(self, key: str, data: Any) -> None:
        if len(self._store) >= self.max_items:
            try:
                self._store.pop(next(iter(self._store)))
            except Exception:
                self._store.clear()
        self._store[key] = (time.time(), data)

DOMAIN_CACHE = TTLCacheSync(ttl_seconds=DOMAIN_CACHE_TTL)
EMAIL_CACHE = TTLCacheSync(ttl_seconds=int(os.getenv("EMAIL_CACHE_TTL", "86400")))

# ==================== DNS & HTTP helpers (wrap blocking libs) ====================

def _resolver() -> dns.resolver.Resolver:
    """
    Use explicit nameservers (can be overridden by DNS_1 / DNS_2 env),
    and set sensible timeout/lifetime. Use configure=False to avoid
    using /etc/resolv.conf if you want explicit control.
    """
    r = dns.resolver.Resolver(configure=False)
    # prefer public resolvers; allow override via env
    dns1 = os.getenv("DNS_1", "1.1.1.1")
    dns2 = os.getenv("DNS_2", "8.8.8.8")
    dns3 = os.getenv("DNS_3", "8.8.4.4")
    dns4 = os.getenv("DNS_4", "1.0.0.1")
    r.nameservers = [dns1, dns2, dns3, dns4]
    # use centralized DNS_TIMEOUT / DNS_LIFETIME
    r.timeout = float(os.getenv("DNS_TIMEOUT", str(DNS_TIMEOUT)))
    r.lifetime = float(os.getenv("DNS_LIFETIME", str(DNS_LIFETIME)))
    # retry SERVFAIL automatically at resolver level
    r.retry_servfail = True
    return r

async def lookup_mx_async(domain: str) -> Tuple[List[str], bool]:
    """
    Resolve MX records for a domain in a synchronous function executed
    via asyncio.to_thread to avoid blocking the event loop.

    Returns: (mx_hosts, implicit_mx_flag)
    - mx_hosts: list of hostnames (strings) — empty list if none or in error cases
    - implicit_mx_flag: True if we used implicit MX logic (i.e., no explicit MX)
    Raises:
    - DNSTimeoutError when resolver timed out after MAX_DNS_RETRIES attempts (hard failure)
    """
    domain = (domain or "").strip().lower().rstrip(".")

    def _do() -> Tuple[List[str], bool]:
        res = _resolver()
        mx_hosts: List[str] = []
        implicit_mx = False

        # We'll attempt up to MAX_DNS_RETRIES (with exponential backoff)
        for attempt in range(MAX_DNS_RETRIES):
            try:
                answers = res.resolve(domain, "MX", lifetime=DNS_TIMEOUT)
                # collect mx records
                mx_hosts = []
                for r in answers:
                    # answer.to_text may give "10 mx.example.com."
                    try:
                        txt = r.to_text()
                        parts = txt.split()
                        if len(parts) == 1:
                            host = parts[0]
                        else:
                            host = parts[-1]
                        host = host.strip().rstrip(".")
                        if host:
                            mx_hosts.append(host)
                    except Exception:
                        try:
                            host = str(r.exchange).rstrip(".")
                            mx_hosts.append(host)
                        except Exception:
                            continue
                # sort by preference if available (dnspython preserves order but ensure stable)
                # If no MX entries, we'll fall through to NoAnswer / implicit MX handling
                if mx_hosts:
                    return list(dict.fromkeys(mx_hosts)), False
                # if answers existed but no mx parsed, treat as no answer to fall back
                raise dns.resolver.NoAnswer()

            except (dns.exception.Timeout, dns.resolver.LifetimeTimeout):
                # DNS request took too long → treat as real timeout
                logger.warning(f"DNS timeout for {domain}, attempt {attempt + 1}/{MAX_DNS_RETRIES}")
                if attempt == MAX_DNS_RETRIES - 1:
                    logger.error(f"DNS timeout for {domain} after {MAX_DNS_RETRIES} retries")
                    # Hard timeout after retries — caller requested to mark these invalid
                    raise DNSTimeoutError(f"DNS timeout for {domain}")
                # exponential backoff
                time.sleep(0.5 * (2 ** attempt))
                continue

            except dns.resolver.NXDOMAIN:
                # Domain doesn't exist -> return empty list; caller may mark invalid
                logger.info(f"NXDOMAIN for {domain}")
                return [], False

            except dns.resolver.NoAnswer:
                # No explicit MX; we'll try implicit MX (A/AAAA) below
                logger.info(f"No MX answer for {domain} (NoAnswer)")
                # try to resolve A/AAAA as implicit MX fallback
                try:
                    ips = []
                    for a in res.resolve(domain, "A", lifetime=DNS_TIMEOUT):
                        ips.append(str(a))
                    for a in res.resolve(domain, "AAAA", lifetime=DNS_TIMEOUT):
                        ips.append(str(a))
                    if ips:
                        # if domain has A/AAAA, treat domain itself as implicit MX
                        logger.info(f"Implicit MX via A/AAAA for {domain}")
                        return [domain], True
                except Exception:
                    # If A/AAAA failed, just return empty MX (caller will deal)
                    pass
                return [], False

            except dns.resolver.NoNameservers as e:
                logger.warning(f"No nameservers response for {domain}: {e}")
                # Try one quick fallback with environment-specified resolvers (if defined)
                try:
                    fallback = dns.resolver.Resolver(configure=False)
                    fallback.nameservers = [ns for ns in (os.getenv("DNS_1"), os.getenv("DNS_2")) if ns]
                    if not fallback.nameservers:
                        return [], False
                    fallback.timeout = DNS_TIMEOUT
                    fallback.lifetime = DNS_LIFETIME
                    answers = fallback.resolve(domain, "MX", lifetime=DNS_TIMEOUT)
                    mx_hosts = []
                    for r in answers:
                        try:
                            txt = r.to_text()
                            parts = txt.split()
                            host = parts[-1] if parts else ""
                            host = host.strip().rstrip(".")
                            if host: mx_hosts.append(host)
                        except Exception:
                            continue
                    if mx_hosts:
                        return list(dict.fromkeys(mx_hosts)), False
                    return [], False
                except Exception as e2:
                    logger.warning(f"Fallback resolver failed for {domain}: {e2}")
                    return [], False

            except dns.exception.DNSException as e:
                # Generic DNS exception (SERVFAIL etc) — treat as non-timeout, non-fatal
                logger.warning(f"Non-timeout DNS exception for {domain}: {e}")
                return [], False

            except Exception as e:
                # Unexpected lower-level exception — do not raise, return empty
                logger.error(f"Unexpected exception resolving MX for {domain}: {e}")
                return [], False

        # If we exit loop without return, treat as empty
        return [], False

    # run blocking resolver in thread pool
    return await asyncio.to_thread(_do)

async def lookup_txt_async(name: str) -> List[str]:
    name = (name or "").strip()
    def _do():
        res = _resolver()
        try:
            answers = res.resolve(name, "TXT", lifetime=DNS_TIMEOUT)
            out = []
            for r in answers:
                try:
                    if hasattr(r, "strings"):
                        parts = [(p.decode(errors="ignore") if isinstance(p, (bytes, bytearray)) else str(p)) for p in r.strings]
                        out.append("".join(parts))
                    else:
                        out.append(r.to_text().strip('"'))
                except Exception:
                    out.append(r.to_text().strip('"'))
            return out
        except Exception:
            return []
    return await asyncio.to_thread(_do)

async def http_head_probe_async(domain: str, timeout: float = 1.0) -> Dict[str, Any]:
    domain = (domain or "").strip()
    def _do():
        url = f"http://{domain}/"
        req = urllib.request.Request(url, method="HEAD")
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                server = r.headers.get("Server", "")
                ct = r.headers.get("Content-Type", "")
                return {"ok": True, "server": server, "content_type": ct, "status": getattr(r, "status", 200)}
        except Exception:
            try:
                with urllib.request.urlopen(f"http://{domain}/", timeout=timeout) as r:
                    server = r.headers.get("Server", "")
                    ct = r.headers.get("Content-Type", "")
                    return {"ok": True, "server": server, "content_type": ct, "status": getattr(r, "status", 200)}
            except Exception as e:
                return {"ok": False, "error": str(e)}
    return await asyncio.to_thread(_do)

async def resolve_a_async(host: str) -> List[str]:
    host = str(host).rstrip(".")
    def _do():
        res = _resolver()
        ips = []
        try:
            for a in res.resolve(host, "A", lifetime=DNS_TIMEOUT):
                ips.append(str(a))
        except Exception:
            pass
        try:
            for a in res.resolve(host, "AAAA", lifetime=DNS_TIMEOUT):
                ips.append(str(a))
        except Exception:
            pass
        return ips
    return await asyncio.to_thread(_do)

async def asn_lookup_cymru_async(ip: str) -> Optional[str]:
    def _do():
        try:
            parts = ip.split(".")
            if len(parts) == 4:
                rev = ".".join(reversed(parts))
                qname = f"{rev}.origin.asn.cymru.com"
            else:
                return None
            res = _resolver()
            answers = res.resolve(qname, "TXT", lifetime=DNS_TIMEOUT)
            for r in answers:
                txt = r.to_text().strip('"')
                if "|" in txt:
                    asn = txt.split("|")[0].strip()
                    if asn:
                        return asn
        except Exception:
            return None
        return None
    return await asyncio.to_thread(_do)

# ==================== SMTP helpers (pooling & probe) ====================

def _make_mailfrom(smtp_from: str) -> str:
    smtp_from = (smtp_from or "").strip()
    try:
        local_host = socket.getfqdn() or "localhost"
    except Exception:
        local_host = "localhost"
    if smtp_from == "" or smtp_from.endswith("@example.com") or smtp_from.endswith("@example.org") or smtp_from.endswith("@example.net"):
        return f"noreply@{local_host}"
    return smtp_from

def _smtp_rcpt_once(mx_host: str, smtp_from: str, rcpt: str, timeout: int = 7) -> Dict[str, Any]:
    res = {"code": None, "message": "", "accepted": None, "perm": False, "temp": False, "retries": 0}
    mx_host = str(mx_host).rstrip(".")
    smtp_from = _make_mailfrom(smtp_from)
    try:
        local_hostname = socket.getfqdn() or "localhost"
    except Exception:
        local_hostname = "localhost"
    try:
        s = smtplib.SMTP(timeout=timeout)
        s.connect(mx_host, 25)
        try:
            s.ehlo()
        except Exception:
            try:
                s.helo(name=local_hostname)
            except Exception:
                pass
        try:
            s.mail(smtp_from)
        except Exception:
            pass

        # Try RCPT up to 2 times only if server returned a 4xx (temporary) on first try.
        code, msg = s.rcpt(rcpt)
        try:
            i_code = int(code) if code is not None else None
        except Exception:
            i_code = None
        msg_text = msg.decode(errors="ignore") if isinstance(msg, (bytes, bytearray)) else str(msg)
        res["code"] = i_code
        res["message"] = msg_text

        if i_code is not None:
            if 200 <= i_code < 300:
                res["accepted"] = True
            elif 400 <= i_code < 500:
                # temporary error — short single retry (small 1s backoff) to handle greylisting
                res["temp"] = True; res["accepted"] = False; res["retries"] = 0
                try:
                    time.sleep(1.0)  # tiny backoff — keeps overall runtime acceptable
                    res["retries"] = 1
                    code2, msg2 = s.rcpt(rcpt)
                    i_code2 = int(code2) if code2 is not None else None
                    msg_text2 = msg2.decode(errors="ignore") if isinstance(msg2, (bytes, bytearray)) else str(msg2)
                    res["code"] = i_code2
                    res["message"] = msg_text2
                    if i_code2 is not None and 200 <= i_code2 < 300:
                        res["accepted"] = True
                        res["temp"] = False
                    elif i_code2 is not None and 500 <= i_code2 < 600:
                        res["perm"] = True; res["accepted"] = False
                except Exception:
                    pass
            elif 500 <= i_code < 600:
                res["perm"] = True; res["accepted"] = False
        else:
            # no numeric code -> conservative temporary
            res["temp"] = True; res["accepted"] = None

        if SMTP_USER_NOT_FOUND_RE.search(res["message"] or ""):
            res["perm"] = True; res["accepted"] = False

        try:
            s.quit()
        except Exception:
            try: s.close()
            except Exception:
                pass

    except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError, socket.timeout, socket.gaierror, ConnectionRefusedError) as e:
        res["message"] = f"connect_error:{e}"
        res["accepted"] = None
        res["temp"] = True
    except Exception as e:
        res["message"] = f"error:{e}"
        res["accepted"] = None
    return res

async def smtp_probe_host_async(host: str, smtp_from: str, rcpt: str) -> Dict[str, Any]:
    return await asyncio.to_thread(_smtp_rcpt_once, host, smtp_from, rcpt, int(os.getenv("SMTP_TIMEOUT", SMTP_TIMEOUT)))

async def probe_domain_async(mx_hosts: List[str], domain: str, smtp_from: str, target_email: str, implicit_mx: bool) -> Dict[str, Any]:
    domain = (domain or "").strip().lower()
    if not mx_hosts:
        return {"smtp_ok": None, "catch_all": None, "hosts": [], "reason": "no mx hosts", "smtp_codes": [], "probe_retries": 0}

    rand_local1 = "noexist_" + ''.join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=10))
    rand_local2 = "noexist2_" + ''.join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=10))
    random_addr1 = f"{rand_local1}@{domain}"
    random_addr2 = f"{rand_local2}@{domain}"
    host = mx_hosts[0]

    # Run target + two random probes concurrently (keeps wall-clock time same)
    target_task = smtp_probe_host_async(host, smtp_from, target_email)
    random_task1 = smtp_probe_host_async(host, smtp_from, random_addr1)
    random_task2 = smtp_probe_host_async(host, smtp_from, random_addr2)
    target_res, random_res1, random_res2 = await asyncio.gather(target_task, random_task1, random_task2)

    host_results = [{"host": host, "target": target_res, "random1": random_res1, "random2": random_res2}]
    smtp_codes = []
    for hr in host_results:
        if hr["target"].get("code"): smtp_codes.append(hr["target"].get("code"))
        if hr["random1"].get("code"): smtp_codes.append(hr["random1"].get("code"))
        if hr["random2"].get("code"): smtp_codes.append(hr["random2"].get("code"))

    target_accept = host_results[0]["target"].get("accepted") is True
    random_accept_any = any(x.get("accepted") is True for x in [host_results[0]["random1"], host_results[0]["random2"]])
    temp_any = any((host_results[0]["target"].get("temp"), host_results[0]["random1"].get("temp"), host_results[0]["random2"].get("temp")))

    smtp_ok = None; catch_all = None; reason = ""
    if target_accept and not random_accept_any:
        smtp_ok = True; catch_all = False; reason = "Target accepted, randoms rejected"
    elif host_results[0]["target"].get("perm") and not random_accept_any:
        smtp_ok = False; catch_all = False; reason = "Hosts permanently rejected target"
    elif target_accept and random_accept_any:
        smtp_ok = None; catch_all = True; reason = "Accepts target AND random (catch-all)"
    elif temp_any:
        smtp_ok = None; catch_all = None; reason = "Temporary errors (4xx/greylisting)"
    else:
        smtp_ok = None; catch_all = None; reason = "No clear acceptance pattern"

    probe_retries = 0
    for hr in host_results:
        probe_retries += (hr["target"].get("retries") or 0)
        probe_retries += (hr["random1"].get("retries") or 0)
        probe_retries += (hr["random2"].get("retries") or 0)

    return {"smtp_ok": smtp_ok, "catch_all": catch_all, "hosts": host_results, "reason": reason, "smtp_codes": smtp_codes, "probe_retries": probe_retries}
# ==================== MX/ASN heuristics ====================

def mx_matches_disposable_keywords(mx_hosts: List[str]) -> bool:
    mxs = ",".join(mx_hosts).lower()
    return any(k in mxs for k in DISPOSABLE_MX_KEYWORDS)

async def asn_fingerprints_for_mx(mx_hosts: List[str]) -> List[str]:
    asns = []
    if not mx_hosts:
        return asns
    ips = await resolve_a_async(mx_hosts[0])
    if not ips:
        return asns
    tasks = [asn_lookup_cymru_async(ip) for ip in ips]
    res = await asyncio.gather(*tasks, return_exceptions=True)
    for r in res:
        if isinstance(r, str) and r:
            asn = r.upper()
            if not asn.startswith("AS"):
                asn = "AS" + asn
            asns.append(asn)
    return list(dict.fromkeys(asns))

# ==================== Validation (single) - combines scoring from old app.py ====================

def _count_chars(local: str) -> Tuple[int,int,int]:
    return (
        sum(1 for c in local if c.isalpha()),
        sum(1 for c in local if c.isdigit()),
        sum(1 for c in local if ord(c) > 127)
    )

async def validate_single_async(email: str, smtp_from: str, db_conn, smtp_probe_flag: bool) -> Dict[str, Any]:
    """
    Full validation with parallel DNS/HTTP/SMTP and scoring logic.
    This mirrors the original app.py scoring but runs network operations concurrently.
    """
    res: Dict[str, Any] = {
        "email": email, "local_part": None, "domain": None,
        "syntax_ok": False,
        "disposable": False, "role_based": False, "free_provider": False,
        "mx_hosts": [], "mx_ok": False, "implicit_mx_record": False,
        "spf": False, "dkim": False, "dmarc": False,
        "smtp_tested": False, "smtp_ok": None, "catch_all": None, "smtp_reason": "",
        "score": None, "final_status": None, "notes": [], "probe_details": None,
        "state": None, "reason": None,
        "smtp_provider": None, "mx_record": None,
        "free": False, "role": False, "accept_all": False, "tag": False,
        "numerical_characters": 0, "alphabetical_characters": 0, "unicode_symbols": 0,
        "mailbox_full": False, "no_reply": False, "secure_email_gateway": False,
        "disposable_signals": {"mx_keyword": False, "asn_match": False, "http_marker": False}
    }

    email_str = (email or "").strip()
    res["email"] = email_str
    if not email_str or not EMAIL_REGEX.match(email_str):
        res["final_status"] = "invalid"; res["syntax_ok"] = False
        res["score"] = 0; res["state"] = "Undeliverable"; res["reason"] = "Invalid format"
        return res
    res["syntax_ok"] = True

    local, domain = email_str.rsplit("@", 1)
    domain = domain.strip().lower()
    res["local_part"] = local; res["domain"] = domain

    alpha, digits, uni = _count_chars(local)
    res["alphabetical_characters"] = alpha
    res["numerical_characters"] = digits
    res["unicode_symbols"] = uni
    res["no_reply"] = local.lower().startswith(("no-reply","noreply"))
    res["role"] = local.lower() in ROLE_LOCALPARTS or any(local.lower().startswith(r + "+") for r in ROLE_LOCALPARTS)
    res["free"] = domain in FREE_PROVIDERS
    res["tag"] = "+" in local

    # Domain-level cache
    cached = DOMAIN_CACHE.get(domain)
    if cached:
        mx_hosts = cached.get("mx_hosts", [])
        implicit_mx = cached.get("implicit_mx", False)
        spf = cached.get("spf", False)
        dkim = cached.get("dkim", False)
        dmarc = cached.get("dmarc", False)
        http_info = cached.get("http_info", {"ok": False})
        asns = cached.get("asns", [])
    else:
        try:
            # create tasks so we can await them exactly once
            mx_task = asyncio.create_task(lookup_mx_async(domain))
            spf_task = asyncio.create_task(lookup_txt_async(domain))
            dkim_task = asyncio.create_task(lookup_txt_async(f"default._domainkey.{domain}"))
            dmarc_task = asyncio.create_task(lookup_txt_async(f"_dmarc.{domain}"))
            http_task = asyncio.create_task(http_head_probe_async(domain, timeout=float(os.getenv("HTTP_PROBE_TIMEOUT", "1.0"))))

            # Wait for MX lookup with a generous timeout (caller-level). If MX causes DNSTimeoutError -> mark invalid
            try:
                mx_hosts, implicit_mx = await asyncio.wait_for(mx_task, timeout=6.0)
            except DNSTimeoutError as e:
                logger.error(f"MX lookup DNSTimeoutError for {domain}: {e}")
                res["final_status"] = "invalid"
                res["syntax_ok"] = True
                res["score"] = 0
                res["state"] = "Undeliverable"
                res["reason"] = f"DNS timeout: {str(e)}"
                res["notes"].append("DNS resolution timeout")
                return res
            except asyncio.TimeoutError:
                # Our overall wait_for timed out -> treat as DNS timeout per requirements
                logger.error(f"DNS batch timeout for domain {domain} (wait_for)")
                res["final_status"] = "invalid"
                res["syntax_ok"] = True
                res["score"] = 0
                res["state"] = "Undeliverable"
                res["reason"] = f"DNS timeout: Domain resolution exceeded 6 seconds"
                res["notes"].append("DNS resolution timeout")
                return res
            except Exception as e:
                # Non-timeout errors while getting MX — treat as no MX and continue
                logger.warning(f"Non-timeout error during MX lookup for {domain}: {e}")
                mx_hosts, implicit_mx = [], False

            # Now await other tasks (they are separate tasks — await once)
            try:
                txts_spf = await spf_task
            except Exception:
                txts_spf = []
            try:
                txts_dkim = await dkim_task
            except Exception:
                txts_dkim = []
            try:
                txts_dmarc = await dmarc_task
            except Exception:
                txts_dmarc = []
            try:
                http_info = await http_task
            except Exception:
                http_info = {"ok": False}

            spf = False; spf_policy = None
            for t in (txts_spf or []):
                tl = t.lower()
                if "v=spf1" in tl:
                    spf = True
                    if "-all" in tl: spf_policy = "-all"
                    elif "~all" in tl: spf_policy = "~all"
                    elif "?all" in tl: spf_policy = "?all"
                    break
            res["spf_policy"] = spf_policy
            dkim = any("v=dkim1" in t.lower() for t in (txts_dkim or []))
            dmarc = any("v=dmarc1" in t.lower() for t in (txts_dmarc or []))

            asns = await asn_fingerprints_for_mx(mx_hosts)

            DOMAIN_CACHE.set(domain, {
                "mx_hosts": mx_hosts, "implicit_mx": implicit_mx,
                "spf": spf, "dkim": dkim, "dmarc": dmarc,
                "http_info": http_info, "asns": asns
            })

        except DNSTimeoutError as e:
            # Defensive: in case any other path raises DNSTimeoutError
            logger.warning(f"DNS timeout for domain {domain}: {e}")
            res["final_status"] = "invalid"
            res["syntax_ok"] = True
            res["score"] = 0
            res["state"] = "Undeliverable"
            res["reason"] = f"DNS timeout or unavailable: {str(e)[:100]}"
            res["notes"].append("DNS resolution failed")
            return res
        except Exception as e:
            # Other unexpected errors during DNS lookup - mark invalid conservatively
            logger.error(f"Unexpected error during DNS lookup for {domain}: {e}")
            res["final_status"] = "invalid"
            res["syntax_ok"] = True
            res["score"] = 0
            res["state"] = "Undeliverable"
            res["reason"] = f"DNS error: {str(e)[:100]}"
            res["notes"].append("DNS lookup error")
            return res

    res["mx_hosts"] = mx_hosts
    res["mx_ok"] = bool(mx_hosts)
    res["implicit_mx_record"] = bool(implicit_mx)
    res["mx_record"] = (mx_hosts[0] if mx_hosts else None)
    res["spf"] = spf; res["dkim"] = dkim; res["dmarc"] = dmarc

    # Lightweight provider tag
    mx_join = ",".join(mx_hosts).lower()
    if "google" in mx_join or "gmail" in mx_join:
        res["smtp_provider"] = "Google"
    elif any(k in mx_join for k in ("outlook","hotmail","office365","protection.outlook")):
        res["smtp_provider"] = "Microsoft"
    elif "yahoo" in mx_join:
        res["smtp_provider"] = "Yahoo"
    else:
        res["smtp_provider"] = "Other"

    # Dynamic disposable detection
    disposable = (domain in DISPOSABLE_DOMAINS)
    mx_keyword = mx_matches_disposable_keywords(mx_hosts)
    asn_match = any(a in DISPOSABLE_ASN_LIST for a in (asns or []))
    http_marker = False
    info = DOMAIN_CACHE.get(domain) if isinstance(DOMAIN_CACHE.get(domain), dict) else {}
    server = (info or {}).get("http_info", {}).get("server","").lower() if info else (http_info.get("server","").lower() if http_info else "")
    if any(k in server for k in ("mailinator","yopmail","temporary","temp-mail","10minutemail")):
        http_marker = True

    res["disposable_signals"] = {"mx_keyword": mx_keyword, "asn_match": asn_match, "http_marker": http_marker}
    if not disposable and (mx_keyword or asn_match or http_marker):
        disposable = True
        res["notes"].append("disposable inferred by MX/ASN/HTTP fingerprint")

    res["disposable"] = disposable

    # Early exit if no MX
    if not mx_hosts:
        res["final_status"] = "invalid"
        res["notes"].append("no MX/A records")
        res["score"] = 0
        res["state"] = "Undeliverable"
        res["reason"] = "No MX or A/AAAA records"
        return res

    # SMTP probing (parallel path)
    if smtp_probe_flag:
        res["smtp_tested"] = True
        probe = await probe_domain_async(mx_hosts, domain, _make_mailfrom(smtp_from), email_str, implicit_mx)
        res["smtp_ok"] = probe["smtp_ok"]
        res["catch_all"] = probe["catch_all"]
        res["smtp_reason"] = probe["reason"]
        res["probe_details"] = probe["hosts"]
        res["smtp_codes"] = probe.get("smtp_codes", []) or []
        res["probe_retries"] = int(probe.get("probe_retries", 0) or 0)
        if probe["reason"]:
            res["notes"].append(probe["reason"])

    # Scoring (kept close to the original app.py logic)
    score = 50
    if res["mx_ok"]:
        score += 12
        if res["implicit_mx_record"]:
            score -= 6
    if res["disposable"]:
        score -= 35
    if res["role"]:
        score -= 10
    if res["free"]:
        score += 5
    if res.get("spf"): score += 4
    if res.get("dkim"): score += 4
    if res.get("dmarc"): score += 4

    if smtp_probe_flag:
        if res["smtp_ok"] is True: score += 25
        elif res["smtp_ok"] is False: score -= 30
        else: score -= 3
        if res["catch_all"]: score -= 15

    pr = PROVIDER_RULES.get(domain)
    if pr:
        mode = pr.get("mode")
        if mode == "trust_mx":
            if res["mx_ok"] and (res["smtp_ok"] is not False): score += 8
            if res["smtp_ok"] is False: score -= 20
        elif mode == "strict":
            if res["smtp_ok"] is True and not res["catch_all"]: score += 12
            elif res["smtp_ok"] is None: score -= 6

    # Guardrails
    if not smtp_probe_flag:
        score = min(score, 60)
    elif res.get("smtp_ok") is None:
        score = min(score, 70)
    if res.get("catch_all"): score = min(score, 60)
    if res.get("smtp_ok") is False: score = min(score, 30)
    if res.get("mx_ok") and res.get("implicit_mx_record"):
        score = min(score, 50)

    score = max(0, min(100, int(score)))
    res["score"] = score

    if res.get("smtp_ok") is True and not res.get("catch_all") and score >= 65:
        final = "valid"
    elif score >= 45:
        final = "risky"
    else:
        final = "invalid"
    res["final_status"] = final
    res["state"] = "Deliverable" if final == "valid" else ("Risky" if final == "risky" else "Undeliverable")
    if res.get("smtp_ok") is True:
        res["reason"] = "ACCEPTED EMAIL"
    elif res.get("smtp_ok") is False:
        res["reason"] = "REJECTED EMAIL"
    else:
        res["reason"] = res["smtp_reason"] or (res["notes"][0] if res["notes"] else "")

    res["accept_all"] = bool(res.get("catch_all"))

    # mailbox full / seg
    if res.get("probe_details"):
        for host in res["probe_details"]:
            tmsg = (host.get("target") or {}).get("message") or ""
            if "mailbox full" in tmsg.lower() or "quota exceeded" in tmsg.lower():
                res["mailbox_full"] = True

    if any(x for x in mx_join.split(",") if any(g in x for g in ("protection","proofpoint","barracuda","mxlogic","emailfilter","email-protection"))):
        res["secure_email_gateway"] = True

    if res["disposable"]:
        res["notes"].append("disposable domain")
    if res["role"]:
        res["notes"].append("role-based mailbox")
    if res["catch_all"]:
        res["notes"].append("catch-all detected")
    if res["smtp_tested"] and res["smtp_ok"] is False:
        res["notes"].append("smtp rejected (permanent)")

    return res

# ==================== Domain-batched bulk processing (optimized) ====================

async def validate_single_optimized(
    email: str,
    smtp_from: str,
    smtp_flag: bool,
    mx_hosts: List[str] = None,
    implicit_mx: bool = False
) -> Dict[str, Any]:
    """
    A lighter-weight per-email function used by domain-batched processing.
    It mirrors important fields from the heavy validate_single_async but is optimized.
    """

    # Check cache first
    email_hash = hashlib.md5(email.encode()).hexdigest()
    cached = EMAIL_CACHE.get(email_hash)
    if cached:
        return cached

    # Reuse the heavier function for full behavior (maintains parity)
    # but pass pre-fetched mx_hosts and implicit flag to avoid redundant DNS.
    res = await validate_single_async(email, smtp_from, None, smtp_flag)
    EMAIL_CACHE.set(email_hash, res)
    return res

async def validate_batch_by_domain(emails: List[str], smtp_from: str, smtp_flag: bool, workers: int) -> List[Dict[str, Any]]:
    # Group by domain
    domain_groups: Dict[str, List[str]] = defaultdict(list)
    for email in emails:
        if "@" in email:
            domain = email.split("@")[1].lower().strip()
            domain_groups[domain].append(email)

    logger.info(f"Processing {len(emails)} emails across {len(domain_groups)} domains")

    all_results = []
    domain_tasks = []
    domain_sem = asyncio.Semaphore(max(1, CPU_CORES * 2))

    async def process_domain_group(domain: str, group_emails: List[str]):
        async with domain_sem:
            try:
                mx_hosts, implicit_mx = await lookup_mx_async(domain)
            except DNSTimeoutError as e:
                # If domain-level MX lookup hard-timed out, mark all emails for that domain as invalid
                logger.error(f"Domain {domain} DNS timeout: {e}. Marking group invalid.")
                return [{
                    "email": eaddr,
                    "final_status": "invalid",
                    "state": "Undeliverable",
                    "reason": f"DNS timeout for domain {domain}",
                    "score": 0,
                    "domain": domain
                } for eaddr in group_emails]
            if not mx_hosts:
                # No MX - mark all invalid
                return [{
                    "email": e,
                    "final_status": "invalid",
                    "state": "Undeliverable",
                    "reason": "No MX records",
                    "score": 0,
                    "domain": domain
                } for e in group_emails]

            DOMAIN_CACHE.set(domain, {
                "mx_hosts": mx_hosts,
                "implicit_mx": implicit_mx
            })

            tasks = [validate_single_optimized(e, smtp_from, smtp_flag, mx_hosts, implicit_mx) for e in group_emails]
            return await asyncio.gather(*tasks)

    for domain, group_emails in domain_groups.items():
        domain_tasks.append(process_domain_group(domain, group_emails))

    results_by_domain = await asyncio.gather(*domain_tasks)
    for domain_results in results_by_domain:
        all_results.extend(domain_results)

    return all_results

async def validate_many_async(emails: List[str], smtp_from: str, smtp_flag: bool, workers: int) -> List[Dict[str, Any]]:
    if not emails:
        return []

    logger.info(f"Starting validation of {len(emails)} emails with {workers} workers (SMTP mode: {SMTP_MODE})")
    start_time = time.time()

    # Use domain-batched processing
    results = await validate_batch_by_domain(emails, smtp_from, smtp_flag, workers)

    elapsed = time.time() - start_time
    rate = len(emails) / elapsed if elapsed > 0 else 0
    logger.info(f"Completed {len(emails)} emails in {elapsed:.2f}s ({rate:.1f} emails/sec)")

    return results

# ==================== CSV helpers ====================

def _strip_email_candidate(value: str) -> str:
    if not value:
        return ""
    return value.strip().strip(EMAIL_STRIP_CHARS).strip()


def _extract_emails_from_value(value: Any) -> List[str]:
    out: List[str] = []
    if value is None:
        return out
    if isinstance(value, bytes):
        try:
            text = value.decode("utf-8", errors="ignore")
        except Exception:
            text = value.decode("latin-1", errors="ignore")
    else:
        text = str(value)
    text = text.strip()
    if not text or "@" not in text:
        return out

    candidate = _strip_email_candidate(text)
    if candidate and EMAIL_REGEX.match(candidate):
        out.append(candidate)
        return out

    for token in EMAIL_TOKEN_SPLIT.split(text):
        token = _strip_email_candidate(token)
        if token and EMAIL_REGEX.match(token):
            out.append(token)
    return out


def _collect_emails_from_rows(rows_iter: Iterable, seen: set[str]) -> List[str]:
    collected: List[str] = []
    for row in rows_iter:
        if row is None:
            continue
        if isinstance(row, (list, tuple)):
            cells = row
        else:
            cells = [row]
        for cell in cells:
            for candidate in _extract_emails_from_value(cell):
                key = candidate.lower()
                if key not in seen:
                    seen.add(key)
                    collected.append(candidate)
    return collected


def _load_from_csv_like(path: str, seen: set[str]) -> List[str]:
    emails: List[str] = []
    try:
        with open(path, newline="", encoding="utf-8", errors="ignore") as f:
            sample = f.read(2048)
            f.seek(0)
            dialect = None
            if sample:
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    pass
            reader = csv.reader(f, dialect or csv.excel)
            emails.extend(_collect_emails_from_rows(reader, seen))
    except Exception as e:
        logger.error(f"Failed to read delimited file {path}: {e}")
    return emails


def _load_from_excel_file(path: str, seen: set[str]) -> List[str]:
    if load_workbook is None:
        logger.warning("openpyxl is not installed; cannot parse Excel file %s", path)
        return []
    emails: List[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            emails.extend(_collect_emails_from_rows(sheet.iter_rows(values_only=True), seen))
        wb.close()
    except Exception as e:
        logger.error(f"Failed to read Excel file {path}: {e}")
    return emails


def _scan_text_file_for_emails(path: str, seen: set[str]) -> List[str]:
    emails: List[str] = []
    try:
        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                emails.extend(_collect_emails_from_rows([line], seen))
    except Exception as e:
        logger.error(f"Failed to scan text file {path}: {e}")
    return emails


CSV_LIKE_EXTENSIONS = {".csv", ".tsv", ".tab", ".txt", ".log", ".dat"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}


def load_emails_from_csv(path: str) -> List[str]:
    return _load_from_csv_like(path, set())


def load_emails_from_file(path: str) -> List[str]:
    seen: set[str] = set()
    emails: List[str] = []
    ext = Path(path).suffix.lower()

    if ext in EXCEL_EXTENSIONS:
        emails.extend(_load_from_excel_file(path, seen))
        if not emails:
            emails.extend(_load_from_csv_like(path, seen))
    else:
        emails.extend(_load_from_csv_like(path, seen))

    if not emails:
        emails.extend(_scan_text_file_for_emails(path, seen))

    return emails

def write_outputs(results: List[Dict[str, Any]], outdir: str) -> None:
    os.makedirs(outdir, exist_ok=True)
    csv_path = os.path.join(outdir, "results.csv")
    json_path = os.path.join(outdir, "results.json")
    if not results:
        logger.info("No results to write."); return

    # CSV header similar to original app.py (concise)
    header = [
        "email","state","reason","final_status","score","syntax_ok",
        "local_part","domain",
        "mx_ok","mx_record","implicit_mx_record","mx_hosts",
        "smtp_tested","smtp_ok","catch_all","smtp_reason",
        "free","role","disposable","accept_all","tag",
        "numerical_characters","alphabetical_characters","unicode_symbols",
        "mailbox_full","no_reply","secure_email_gateway",
        "smtp_provider","notes","probe_details",
        "disposable_mx_keyword","disposable_asn_match","disposable_http_marker",
        "smtp_codes","spf_policy","probe_retries"
    ]

    rows = []
    for r in results:
        row = {}
        for h in header:
            if h == "mx_hosts":
                row[h] = ",".join(r.get("mx_hosts") or [])
            elif h == "probe_details":
                pd = r.get("probe_details")
                if not pd:
                    row[h] = ""
                else:
                    parts = []
                    for hr in pd:
                        host = hr.get("host")
                        t = hr.get("target") or {}
                        # support legacy single 'random' key or new 'random1'/'random2'
                        r1 = hr.get("random") or hr.get("random1") or {}
                        r2 = hr.get("random2") or {}
                        # format codes & accepted flags defensively
                        tc = t.get("code") or ""
                        ta = int(bool(t.get("accepted")))
                        r1c = r1.get("code") or ""
                        r1a = int(bool(r1.get("accepted")))
                        r2c = r2.get("code") or ""
                        r2a = int(bool(r2.get("accepted")))
                        parts.append(
                            f"{host}:T({tc},{ta})R1({r1c},{r1a})R2({r2c},{r2a})"
                        )
                    row[h] = " | ".join(parts)
            elif h == "notes":
                row[h] = " ; ".join(r.get("notes") or [])
            elif h == "disposable_mx_keyword":
                row[h] = int(bool((r.get("disposable_signals") or {}).get("mx_keyword")))
            elif h == "disposable_asn_match":
                row[h] = int(bool((r.get("disposable_signals") or {}).get("asn_match")))
            elif h == "disposable_http_marker":
                row[h] = int(bool((r.get("disposable_signals") or {}).get("http_marker")))
            elif h == "smtp_codes":
                sc = r.get("smtp_codes")
                if not sc:
                    # try to extract codes from probe_details list
                    pd = r.get("probe_details") or []
                    scs = []
                    for hr in pd:
                        # host may include target/random keys with codes
                        t = hr.get("target") or {}
                        if t.get("code"): scs.append(t.get("code"))
                        r1 = hr.get("random") or hr.get("random1") or {}
                        if r1.get("code"): scs.append(r1.get("code"))
                        r2 = hr.get("random2") or {}
                        if r2.get("code"): scs.append(r2.get("code"))
                    sc = scs
                row[h] = ",".join(str(x) for x in (sc or []))
            elif h == "spf_policy":
                row[h] = r.get("spf_policy") or ""
            elif h == "probe_retries":
                row[h] = int(r.get("probe_retries", 0))
            else:
                row[h] = r.get(h) if h in r else ""
        rows.append(row)

    try:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=header); w.writeheader(); w.writerows(rows)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"results": results}, f, indent=2, ensure_ascii=False)
        logger.info(f"Wrote CSV -> {csv_path}")
        logger.info(f"Wrote JSON -> {json_path}")
    except Exception as e:
        logger.error(f"Failed to write outputs: {e}")

# ==================== CLI Entrypoint ====================

def run_file(input_file: str, outdir: str, smtp_flag: bool, smtp_from: str, workers: int) -> None:
    emails = load_emails_from_file(input_file)
    if not emails:
        write_outputs([], outdir); return
    # Cap workers sensibly
    workers = max(1, min(workers or DEFAULT_WORKERS, 1000))
    results = asyncio.run(validate_many_async(emails, smtp_from, smtp_flag, workers))
    write_outputs(results, outdir)

def parse_args():
    import argparse
    p = argparse.ArgumentParser(description="Optimized async email verifier")
    p.add_argument("input_file")
    p.add_argument("--outdir", default="results")
    p.add_argument("--smtp", dest="smtp", action="store_true", default=True)
    p.add_argument("--no-smtp", dest="smtp", action="store_false")
    p.add_argument("--smtp-from", default="jeet8patel1970@gmail.com")
    p.add_argument("--smtp-mode", choices=["fast", "balanced", "thorough"], default="balanced")
    p.add_argument("--workers", type=int, default=DEFAULT_WORKERS)
    return p.parse_args()

if __name__ == "__main__":
    args = parse_args()
    os.environ["SMTP_MODE"] = args.smtp_mode
    SMTP_MODE = args.smtp_mode
    SMTP_TIMEOUT = SMTP_TIMEOUTS.get(SMTP_MODE, SMTP_TIMEOUTS["balanced"])
    run_file(args.input_file, args.outdir, args.smtp, args.smtp_from, args.workers)
